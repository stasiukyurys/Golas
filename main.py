import os
import openpyxl as opx
import operator
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment, Font


a_month = ('январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
           'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь')


class Payer:
    def __init__(self,
                 name,  # наименование плательщика
                 request_date,  # дата запроса
                 number,  # номер теража
                 request_sum,  # сумма запроса
                 total_request_sum,  # итого сумма запроса
                 listed,  # перечисленно
                 total_next_month,  # итого на начало следующего месяца
                 total_first_month,  # итого на начало месяца
                 month,  # месяц
                 comment):  # комментарий
        self.name = name
        self.request_date = request_date
        self.number = number
        self.request_sum = request_sum
        self.total_request_sum = total_request_sum
        self.listed = listed
        self.total_next_month = total_next_month
        self.total_first_month = total_first_month
        self.month = month
        self.comment = comment


def main():
    workbook = opx.load_workbook(file(), read_only=True, data_only=True)
    page1_sheet = workbook.worksheets[0]
    # берем данные для class <Payer> с страницы <Page1>
    i = 0
    a = 4
    global s_Payer
    s_Payer = []
    try:
        for row in page1_sheet.rows:
            i += 1
            for s_month in a_month:
                if row[2].value == s_month:
                    y = 1
                    s = ''
                    while s != 'итого':
                        if page1_sheet.cell(row=i + y, column=3).value is not None:
                            s = page1_sheet.cell(row=i + y, column=3).value
                            s = s.lower()
                            s = s.strip()
                            if s != 'итого':
                                while a < 13:
                                    if page1_sheet.cell(row=i + y, column=a).value is not None:
                                        name = page1_sheet.cell(row=i + y, column=3).value  # наименование плательщика
                                        request_date = page1_sheet.cell(row=i + 1, column=a).value  # дата запроса
                                        number = page1_sheet.cell(row=i, column=a).value  # номер теража
                                        request_sum = page1_sheet.cell(row=i + y, column=a).value  # сумма запроса
                                        total_request_sum = page1_sheet.cell(row=i + y,
                                                                             column=13).value  # итого сумма запроса
                                        listed = page1_sheet.cell(row=i + y, column=14).value  # перечисленно
                                        total_next_month = page1_sheet.cell(row=i + y,
                                                                            column=15).value  # итого на начало следующего месяца
                                        total_first_month = page1_sheet.cell(row=i + y,
                                                                             column=2).value  # итого на начало месяца
                                        month = s_month  # месяц
                                        comment = page1_sheet.cell(row=i + y, column=16).value  # комментарий
                                        s_Payer.append(Payer(name, request_date, number, request_sum, total_request_sum,
                                                             listed, total_next_month, total_first_month, month, comment))
                                        print(str(month)
                                              + '   ' + str(name)
                                              + '   ' + str(request_date)
                                              + '   ' + str(number)
                                              + '   ' + str(request_sum)
                                              + '   ' + str(total_request_sum)
                                              + '   ' + str(listed)
                                              + '   ' + str(total_next_month)
                                              + '   ' + str(total_first_month)
                                              + '   ' + str(comment))
                                    a += 1
                        y += 1
                        a = 4
    except:
        pass
    s_Payer.sort(key=operator.attrgetter('name'))
    excel_withhold()
    workbook.close()


def file():
    directory = './'
    files = os.listdir(directory)
    try:
        excel_file = next(filter(lambda x: x.endswith('.xlsx'), files))
        print('Открыт файл: ' + directory + excel_file)
        return excel_file
    except:
        print('Не могу найти файл *.xlsx или к нему нет доступа')
        print('Для завершения работы нажмите <Enter>')
        input()
        exit()


def excel_withhold():
    # определяем стили
    font = Font(name='Calibri',
                size=12,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
    fill = PatternFill(fill_type='solid', start_color='c1c1c1', end_color='c2c2c2')
    align_center = Alignment(horizontal='center',
                             vertical='bottom',
                             text_rotation=0,
                             wrap_text=False,
                             shrink_to_fit=False,
                             indent=0)
    align_left = Alignment(horizontal='left',
                           vertical='bottom',
                           text_rotation=0,
                           wrap_text=False,
                           shrink_to_fit=False,
                           indent=0)
    # текущее время
    today = datetime.today()
    today = today.strftime('%d.%m.%Y')
    # объект
    wb = opx.Workbook()
    # активный лист
    ws = wb.active
    # название страницы
    ws.title = 'Отчет'
    # значение ячейки
    ws['A1'] = 'Отчет создан: ' + str(today)
    # данные для строк
    rows = [['Месяц',
             'Наименование плательщика',
             'Дата запроса',
             'Номер теража',
             'Сумма запроса',
             'Итого запрос',
             'Нач. месяца',
             'След. месяца',
             'Перечисленно',
             'Комментарий'
             ]]
    # циклом записываем данные
    for row in rows:
        ws.append(row)
    i = 0
    d = 0
    while i < len(s_Payer):
        if ws.cell(row=i + 2 + d, column=2).value != s_Payer[i].name:
            d += 3
        if ws.cell(row=i + 2 + d, column=1).value != s_Payer[i].month:
            ws.cell(row=i + 3 + d, column=6).value = s_Payer[i].total_request_sum
            ws.cell(row=i + 3 + d, column=6).font = font
            ws.cell(row=i + 3 + d, column=7).value = s_Payer[i].total_first_month
            ws.cell(row=i + 3 + d, column=7).font = font
            ws.cell(row=i + 3 + d, column=8).value = s_Payer[i].total_next_month
            ws.cell(row=i + 3 + d, column=8).font = font
            ws.cell(row=i + 3 + d, column=9).value = s_Payer[i].listed
            ws.cell(row=i + 3 + d, column=9).font = font
            ws.cell(row=i + 3 + d, column=10).value = s_Payer[i].comment
            ws.cell(row=i + 3 + d, column=10).font = font

        ws.cell(row=i + 3 + d, column=1).value = s_Payer[i].month
        ws.cell(row=i + 3 + d, column=2).value = s_Payer[i].name
        ws.cell(row=i + 3 + d, column=3).value = s_Payer[i].request_date
        ws.cell(row=i + 3 + d, column=4).value = s_Payer[i].number
        ws.cell(row=i + 3 + d, column=5).value = s_Payer[i].request_sum
        i += 1

    # раскрвшивание фона для заголовков
    ws['A2'].fill = fill
    ws['B2'].fill = fill
    ws['C2'].fill = fill
    ws['D2'].fill = fill
    ws['E2'].fill = fill
    ws['F2'].fill = fill
    ws['G2'].fill = fill
    ws['H2'].fill = fill
    ws['I2'].fill = fill
    ws['J2'].fill = fill
    # вручную устанавливаем высоту первой строки
    rd = ws.row_dimensions[2]
    rd.height = 30
    # вручную устанавливаем ширину столбцов
    cd = ws.column_dimensions['A']
    cd.width = 10
    cd = ws.column_dimensions['B']
    cd.width = 30
    cd = ws.column_dimensions['C']
    cd.width = 14
    cd = ws.column_dimensions['D']
    cd.width = 14
    cd = ws.column_dimensions['E']
    cd.width = 14
    cd = ws.column_dimensions['F']
    cd.width = 14
    cd = ws.column_dimensions['G']
    cd.width = 14
    cd = ws.column_dimensions['H']
    cd.width = 14
    cd = ws.column_dimensions['I']
    cd.width = 14
    cd = ws.column_dimensions['J']
    cd.width = 30
    # увеличиваем все строки по высоте
    max_row = ws.max_row
    i = 3
    while i <= max_row:
        rd = ws.row_dimensions[i]
        rd.height = 14
        i += 1
    # сетка + выравнивание
    for cellObj in ws['A2:J2']:
        for cell in cellObj:
            ws[cell.coordinate].alignment = align_center
    # выравнивание столбца
    for cellObj in ws['A2:A3']:
        for cell in cellObj:
            ws[cell.coordinate].alignment = align_left
    # сохранение файла в текущую директорию
    wb.save('report_' + today + '.xlsx')


main()
